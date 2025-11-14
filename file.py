from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===== KHỞI TẠO =====
wb = Workbook()
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
header_font = Font(bold=True, color="000000")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ===== SHEET 1: Tổng quan =====
ws1 = wb.active
ws1.title = "Tổng quan"

overview_data = [
    ["Dự án:", "Minh Khang"],
    ["Tháng:", "11/2025"],
    ["Số sprint:", "2"],
    ["Thời gian mỗi sprint:", "2 tuần"],
    ["Tổng số lần deploy:", "4"],
    ["Ghi chú:", "Mỗi tuần deploy 1 lần, FE & BE cùng review task trước release."]
]

for row in overview_data:
    ws1.append(row)

ws1.column_dimensions["A"].width = 25
ws1.column_dimensions["B"].width = 50
for cell in ws1["A"]:
    cell.font = Font(bold=True)
for row in ws1.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="center")

# ===== HÀM TẠO BẢNG TASK =====
def create_task_table(ws, sprint_name, week_name, tasks):
    ws.append([])
    ws.append([f"{sprint_name} - {week_name}"])
    ws.append(["Ngày Deploy", "Hạng mục", "Mô tả", "FE", "BE",
               "Ưu tiên", "Tiến độ (%)", "Ghi chú"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for t in tasks:
        ws.append(t)
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.append([])

# ===== SHEET 2: Sprint 1 =====
ws2 = wb.create_sheet("Sprint 1")

tasks_week1 = [
    ["08/11/2025", "Quản lý người dùng",
     "CRUD user, phân quyền role, API login/logout, UI danh sách user", "A", "B", "Cao", "80%", "Cần review UI table"],
    ["08/11/2025", "Auth middleware",
     "Thêm middleware xác thực JWT và refresh token", "A", "B", "Trung bình", "60%", ""],
]

tasks_week2 = [
    ["14/11/2025", "Notification realtime",
     "Tích hợp WebSocket hiển thị thông báo real-time", "A", "B", "Cao", "50%", "Đang test kết nối WS"],
    ["14/11/2025", "Push noti API",
     "Tạo endpoint gửi thông báo cho người dùng", "A", "B", "Thấp", "30%", ""],
]

create_task_table(ws2, "Sprint 1", "Tuần 1", tasks_week1)
create_task_table(ws2, "Sprint 1", "Tuần 2", tasks_week2)

for col, width in zip(["A", "B", "C", "D", "E", "F", "G", "H"],
                      [14, 25, 50, 10, 10, 12, 12, 25]):
    ws2.column_dimensions[col].width = width

# ===== SHEET 3: Sprint 2 =====
ws3 = wb.create_sheet("Sprint 2")

tasks_week3 = [
    ["22/11/2025", "Quản lý giao dịch",
     "Xử lý thanh toán, hiển thị lịch sử đơn hàng", "A", "B", "Cao", "40%", ""],
    ["22/11/2025", "API thống kê đơn hàng",
     "Trả về dữ liệu thống kê đơn hàng theo tháng", "A", "B", "Trung bình", "50%", "Cần kiểm thử hiệu năng"],
]

tasks_week4 = [
    ["29/11/2025", "Dashboard tổng hợp",
     "Xây dựng dashboard hiển thị biểu đồ doanh thu và lượt truy cập", "A", "B", "Cao", "30%", "Đang thiết kế UI"],
    ["29/11/2025", "Tối ưu query backend",
     "Giảm thời gian truy vấn bằng caching và index DB", "A", "B", "Cao", "20%", ""],
]

create_task_table(ws3, "Sprint 2", "Tuần 3", tasks_week3)
create_task_table(ws3, "Sprint 2", "Tuần 4", tasks_week4)

for col, width in zip(["A", "B", "C", "D", "E", "F", "G", "H"],
                      [14, 25, 50, 10, 10, 12, 12, 25]):
    ws3.column_dimensions[col].width = width

# ===== LƯU FILE =====
wb.save("task_plan_minh_khang_thang_11_2025.xlsx")
print("✅ Đã tạo file: task_plan_minh_khang_thang_11_2025.xlsx")
